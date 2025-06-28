import os
import openpyxl
from flask import Flask, render_template, request, redirect, url_for
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['ALLOWED_EXTENSIONS'] = {'xlsx'}

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in app.config['ALLOWED_EXTENSIONS']

def format_12h(time_str):
    """Convert 24-hour time string to 12-hour format with AM/PM"""
    if time_str == '-' or not time_str:
        return time_str
    try:
        time_obj = datetime.strptime(time_str, "%H:%M")
        return time_obj.strftime("%I:%M %p").lstrip('0')
    except:
        return time_str

def is_sunday(date_str):
    """Check if a date string (YYYY-MM-DD) is a Sunday"""
    try:
        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
        return date_obj.weekday() == 6  # Sunday is weekday 6
    except:
        return False

def calculate_working_hours(time_in, time_out):
    """Calculate working hours based on actual time in/out"""
    if time_in == '-' or time_out == '-':
        return 0
    try:
        time_in_obj = datetime.strptime(time_in, "%I:%M %p")
        time_out_obj = datetime.strptime(time_out, "%I:%M %p")
        
        # Special handling for 4PM-10PM shift
        if time_in_obj.hour >= 16 and time_out_obj.hour <= 22:  # 4PM-10PM
            # If time out is before time in (unlikely for this shift), add a day
            if time_out_obj < time_in_obj:
                time_out_obj += timedelta(days=1)
        
        # Regular handling for other shifts
        elif time_out_obj < time_in_obj:
            time_out_obj += timedelta(days=1)
        
        total_minutes = (time_out_obj - time_in_obj).total_seconds() / 60
        return total_minutes / 60  # Convert to hours
    except:
        return 0

def is_shifting_employee(department):
    """Check if employee is shifting based on department value"""
    return str(department).strip().lower() == "liveseller"

def get_shift_schedule(time_in_str):
    """Determine the shift schedule based on time-in with precise cutoffs"""
    try:
        time_in = datetime.strptime(time_in_str, "%I:%M %p")
        hour = time_in.hour
        minute = time_in.minute
        total_minutes = hour * 60 + minute
        
        # Define shift boundaries in minutes since midnight
        morning_cutoff = 9 * 60    # 9:00 AM
        afternoon_cutoff = 13 * 60  # 1:00 PM
        early_night_cutoff = 16 * 60  # 4:00 PM (new cutoff for 4PM-10PM shift)
        night_cutoff = 17 * 60      # 5:00 PM
        late_night_cutoff = 19 * 60 # 7:00 PM
        
        # Morning shift (9AM-5PM)
        if total_minutes < morning_cutoff + 10:  # Before 9:10 AM
            return ('9:00 AM', '5:00 PM', 'morning', morning_cutoff)
        # Afternoon shift (1PM-9PM)
        elif total_minutes < afternoon_cutoff + 10:  # Before 1:10 PM
            return ('1:00 PM', '9:00 PM', 'afternoon', afternoon_cutoff)
        # Early Night shift (4PM-10PM) - NEW SHIFT
        elif total_minutes < early_night_cutoff + 10:  # Before 4:10 PM
            return ('4:00 PM', '10:00 PM', 'night', early_night_cutoff)
        # Night shift (5PM-1AM)
        elif total_minutes < night_cutoff + 10:  # Before 5:10 PM
            return ('5:00 PM', '1:00 AM', 'night', night_cutoff)
        # Late Night shift (7PM-1AM)
        else:  # 5:10 PM onwards
            return ('7:00 PM', '1:00 AM', 'night', late_night_cutoff)
    except:
        return (None, None, 'day', 7*60)  # Default to 7AM for day shift

def calculate_late_minutes(time_in, is_shifting):
    """Calculate minutes late with proper grace periods"""
    if time_in == '-':
        return 0
    
    try:
        time_obj = datetime.strptime(time_in, "%I:%M %p")
        total_minutes = time_obj.hour * 60 + time_obj.minute
        
        if is_shifting:
            _, _, _, cutoff = get_shift_schedule(time_in)
            # 10 minute grace period for all shifting employees
            late_minutes = total_minutes - (cutoff + 10)
        else:
            # Regular employees (7AM cutoff with 10 min grace)
            late_minutes = total_minutes - (7*60 + 10)
        
        return max(0, int(late_minutes))
    except:
        return 0

def calculate_overtime_minutes(time_out, is_shifting):
    """Calculate overtime minutes (regular employees only)"""
    if time_out == '-' or is_shifting:
        return 0
    
    try:
        time_out_obj = datetime.strptime(time_out, "%I:%M %p")
        total_minutes = time_out_obj.hour * 60 + time_out_obj.minute
        end_time = 17 * 60  # 5:00 PM
        
        # Handle overnight
        if time_out_obj.hour < 12:  # If AM time
            end_time = 29 * 60  # 5:00 AM next day
        
        if total_minutes <= end_time + 14:  # Before 5:15 PM
            return 0
        
        ot_minutes = total_minutes - end_time
        
        if total_minutes <= end_time + 60:  # 5:15PM-6:00PM
            return int(ot_minutes)
        elif total_minutes <= end_time + 74:  # 6:00PM-6:14PM
            return 60
        else:  # 6:15PM onwards
            return int(60 + (total_minutes - (end_time + 75)))
    except:
        return 0

def process_time_records(raw_times, next_day_times, is_shifting):
    """Process time records with accurate shift detection and proper time in/out handling"""
    if not raw_times and not next_day_times:
        return None
    
    fmt = "%H:%M"
    times = []
    for t in raw_times:
        try:
            time_obj = datetime.strptime(t, fmt)
            if not (1 <= time_obj.hour < 5):  # Filter 1AM-5AM
                times.append(time_obj)
        except:
            continue
    
    next_day_early = []
    for t in next_day_times:
        try:
            t_obj = datetime.strptime(t, fmt) if isinstance(t, str) else t
            if t_obj.hour < 6:  # Before 6AM is next day
                next_day_early.append(t_obj)
        except:
            continue
    
    all_times = sorted(times + next_day_early)
    
    if not all_times:
        return None
    
    # Initialize result with default values
    result = {
        'time_in': '-',
        'time_out': '-',
        'shift_type': 'day',
        'hours_worked': 0,
        'late_minutes': 0,
        'overtime_minutes': 0,
        'notes': 'No time records',
        'status': 'absent'
    }
    
    # ===== SHIFTING EMPLOYEES (LIVESELLERS) =====
    if is_shifting:
        # Special handling for Livesellers with cross-day shifts
        if next_day_early and len(times) >= 1:
            # For Livesellers, we need to properly pair the time out from previous day (next_day_early)
            # with the time in from current day (times)
            
            # Get all PM times from current day (potential time ins)
            pm_times = [t for t in times if t.hour >= 12]
            
            if pm_times:
                # If there are PM times, use the earliest as time in
                time_in = min(pm_times)
                time_out = None
                
                # Find matching time out (should be in next_day_early)
                if next_day_early:
                    time_out = min(next_day_early)
                
                if time_out:
                    time_in_str = format_12h(time_in.strftime(fmt))
                    time_out_str = format_12h(time_out.strftime(fmt))
                    
                    shift_start, shift_end, shift_type, _ = get_shift_schedule(time_in_str)
                    
                    result.update({
                        'time_in': time_in_str,
                        'time_out': time_out_str,
                        'shift_type': shift_type,
                        'hours_worked': calculate_working_hours(time_in_str, time_out_str),
                        'late_minutes': calculate_late_minutes(time_in_str, is_shifting),
                        'notes': f'{shift_type.replace("_", " ").title()} shift ({shift_start}-{shift_end})',
                        'status': 'present'
                    })
                    return result
            
            # Handle case where we have AM times in current day (regular shift)
            am_times = [t for t in times if t.hour < 12]
            if am_times and len(times) >= 2:
                # Regular shift (AM time in, PM time out)
                time_in = min(am_times)
                time_out = max([t for t in times if t > time_in])
                
                time_in_str = format_12h(time_in.strftime(fmt))
                time_out_str = format_12h(time_out.strftime(fmt))
                
                shift_start, shift_end, shift_type, _ = get_shift_schedule(time_in_str)
                
                result.update({
                    'time_in': time_in_str,
                    'time_out': time_out_str,
                    'shift_type': shift_type,
                    'hours_worked': calculate_working_hours(time_in_str, time_out_str),
                    'late_minutes': calculate_late_minutes(time_in_str, is_shifting),
                    'notes': f'{shift_type.replace("_", " ").title()} shift ({shift_start}-{shift_end})',
                    'status': 'present'
                })
                return result
        
        # Handle single time entry cases
        if len(all_times) == 1:
            time_str = format_12h(all_times[0].strftime(fmt))
            
            # For Livesellers, if the single time is between 12AM-6AM, it's likely the time OUT
            if 0 <= all_times[0].hour < 6:
                result.update({
                    'time_in': '-',
                    'time_out': time_str,
                    'status': 'no_time_in',
                    'notes': 'Missing time in (Liveseller single AM entry treated as time out)'
                })
            else:
                if all_times[0].hour < 12:  # AM time - should be time in
                    result.update({
                        'time_in': time_str,
                        'time_out': '-',
                        'status': 'no_time_out',
                        'notes': 'Missing time out'
                    })
                else:  # PM time - should be time out
                    result.update({
                        'time_in': '-',
                        'time_out': time_str,
                        'status': 'no_time_in',
                        'notes': 'Missing time in'
                    })
            return result
    
    # Rest of the regular employee logic remains the same
    # ===== REGULAR EMPLOYEES =====
    if len(all_times) >= 2:
        # For regular employees, AM should be time in, PM should be time out
        am_times = [t for t in all_times if t.hour < 12]
        pm_times = [t for t in all_times if t.hour >= 12]
        
        if am_times and pm_times:
            time_in = min(am_times)
            time_out = max(pm_times)
        else:
            time_in = min(all_times)
            time_out = max(all_times)
        
        time_in_str = format_12h(time_in.strftime(fmt))
        time_out_str = format_12h(time_out.strftime(fmt))
        
        result.update({
            'time_in': time_in_str,
            'time_out': time_out_str,
            'shift_type': 'day',
            'hours_worked': calculate_working_hours(time_in_str, time_out_str),
            'late_minutes': calculate_late_minutes(time_in_str, is_shifting),
            'overtime_minutes': calculate_overtime_minutes(time_out_str, is_shifting),
            'notes': 'Regular schedule (7AM-5PM)',
            'status': 'present'
        })
    elif all_times:
        time_str = format_12h(all_times[0].strftime(fmt))
        if all_times[0].hour < 12:  # AM time - should be time in
            result.update({
                'time_in': time_str,
                'time_out': '-',
                'status': 'no_time_out',
                'notes': 'Missing time out'
            })
        else:  # PM time - should be time out
            result.update({
                'time_in': '-',
                'time_out': time_str,
                'status': 'no_time_in',
                'notes': 'Missing time in'
            })
    
    return result
    
def process_attendance(file_path):
    """Process the Excel attendance file with department detection"""
    try:
        wb = openpyxl.load_workbook(file_path)
        sheet = wb.active
        
        employees = []
        departments = []
        attendance_data = []
        
        for row in range(5, sheet.max_row + 1, 3):
            employee_name = sheet[f'K{row}'].value
            department = sheet[f'U{row}'].value if sheet[f'U{row}'].value else ""
            
            if not employee_name:
                continue
                
            is_shifting = is_shifting_employee(department)
            time_data = {}
            
            for col in range(1, 32):  # Columns A to AE (31 days)
                current_cell = sheet.cell(row=row+1, column=col)
                next_cell = sheet.cell(row=row+1, column=col+1) if col + 1 <= 31 else None
                
                raw_times = []
                if current_cell.value:
                    raw_times = [t.strip() for t in str(current_cell.value).split('\n') if t.strip()]
                
                next_day_times = []
                if next_cell and next_cell.value:
                    next_day_logs = [t.strip() for t in str(next_cell.value).split('\n') if t.strip()]
                    for time_str in next_day_logs:
                        try:
                            time = datetime.strptime(time_str, "%H:%M")
                            next_day_times.append(time)
                        except:
                            continue
                
                processed_times = process_time_records(raw_times, next_day_times, is_shifting)
                if processed_times:
                    time_data[col] = processed_times
            
            employees.append(employee_name)
            departments.append(department)
            attendance_data.append(time_data)
        
        return employees, departments, attendance_data
    
    except Exception as e:
        print(f"Error processing file: {e}")
        return [], [], []

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        if 'file' not in request.files:
            return redirect(request.url)
        
        file = request.files['file']
        if file.filename == '':
            return redirect(request.url)
        
        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            employees, departments, attendance_data = process_attendance(filepath)
            employee_records = []
            current_month = datetime.now().month
            current_year = datetime.now().year
            
            for emp, dept, att in zip(employees, departments, attendance_data):
                is_shifting = is_shifting_employee(dept)
                emp_data = {
                    'name': emp,
                    'department': dept,
                    'days': [],
                    'total_hours': 0,
                    'present_days': 0,
                    'total_late_minutes': 0,
                    'total_overtime_minutes': 0,
                    'is_shifting': is_shifting
                }
                
                for day in range(1, 32):
                    date_str = f"{current_year}-{current_month:02d}-{day:02d}"
                    try:
                        date_obj = datetime.strptime(date_str, "%Y-%m-%d")
                        day_name = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'][date_obj.weekday()]
                        is_sunday = date_obj.weekday() == 6
                    except:
                        day_name = ''
                        is_sunday = False
                    
                    # Handle day off status for Livesellers
                    if is_sunday:
                        if is_shifting:
                            # For Livesellers, Sunday is working day only if they have attendance
                            if day in att:
                                day_data = att[day]
                                status = 'present' if day_data.get('status') != 'absent' else 'day_off'
                            else:
                                status = 'day_off'
                        else:
                            status = 'day_off'
                    else:
                        status = 'absent' if day not in att else att[day].get('status', 'present')
                    
                    if day in att and status != 'day_off':
                        day_data = att[day]
                        hours = float(day_data.get('hours_worked', 0))
                        late = day_data.get('late_minutes', 0)
                        overtime = day_data.get('overtime_minutes', 0)
                        
                        emp_data['days'].append({
                            'day': day,
                            'date': date_str,
                            'day_name': day_name,
                            'is_sunday': is_sunday,
                            'time_in': day_data.get('time_in', '-'),
                            'time_out': day_data.get('time_out', '-'),
                            'hours': f"{hours:.2f}",
                            'late_minutes': late,
                            'overtime_minutes': overtime,
                            'shift_type': day_data.get('shift_type', 'regular'),
                            'status': status,
                            'notes': day_data.get('notes', '')
                        })
                        
                        if status == 'present':
                            emp_data['total_hours'] += hours
                            emp_data['present_days'] += 1
                            emp_data['total_late_minutes'] += late
                            emp_data['total_overtime_minutes'] += overtime
                    else:
                        emp_data['days'].append({
                            'day': day,
                            'date': date_str,
                            'day_name': day_name,
                            'is_sunday': is_sunday,
                            'time_in': '-',
                            'time_out': '-',
                            'hours': '0.00',
                            'late_minutes': 0,
                            'overtime_minutes': 0,
                            'shift_type': '',
                            'status': status,
                            'notes': 'Sunday - Day Off' if is_sunday else 'Absent'
                        })
                
                emp_data['total_hours'] = f"{emp_data['total_hours']:.2f}"
                try:
                    avg = float(emp_data['total_hours']) / emp_data['present_days'] if emp_data['present_days'] > 0 else 0
                    emp_data['average_hours'] = f"{avg:.2f}"
                except:
                    emp_data['average_hours'] = '0.00'
                
                employee_records.append(emp_data)
            
            return render_template('report.html', 
                                employees=employee_records,
                                filename=filename,
                                report_date=datetime.now().strftime("%Y-%m-%d"),
                                month_year=f"{datetime.now().strftime('%B %Y')}")
    
    return render_template('upload.html')

if __name__ == '__main__':
    app.run(debug=True)